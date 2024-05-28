import re
from time import sleep
import logging

import openpyxl
from Utilities.all import pId, error_42, error_06, error_04, error_07, ConfigReset, AdmStateEA
from base.Atlas.atlas import SessionAPI

console_handler = logging.StreamHandler()

logging.basicConfig(format='%(asctime)s - %(levelname)s - %(module)s - %(message)s',
                    datefmt='%m.%d.%Y %H:%M:%S',
                    level=logging.INFO)

deal_params = []


def get_value(lists, key):
    """
    Метод, который возвращает value из словаря по ключу
    """
    for k, v in lists.items():
        if k == key:
            return str(v)


def only_numbers(parameter):
    """
    Проверка параметров с числовыми данными.
    На проверку передается строка с буквами.
    """
    with open('TEST-T152_result.txt', 'a') as file:
        logging.info(f'Передача параметру {parameter} текста "asdфыв"')
        req = t.set_param("{}_{}_{}".format(testing_devcls, testing_slot, parameter), "asdфыв")["RACK"]["PARAM"]["@Msg"]
        if req == '':
            file.write(f'Вместо [ {error_42} ] {parameter} принял строку "asdфыв"\n')
        elif req == error_42:
            logging.info(f'Проверка передачи параметру {parameter} текста "asdфыв" прошла успешно')
        else:
            file.write(f'При передаче строки "asdфыв" параметру {parameter} вместо [ {error_42} ] пришел ответ [ {req} ]\n')


def send_str(parameter, data_type):
    """
    Проверка параметров со строковыми данными.
    На проверку передаются строка длиной 150 символов. (для параметров, которые ожидают на вход строку).
    На проверку передается строка длиной 5 символов. (для параметров, которые ожидают на вход числовые значения)
    """
    with open('TEST-T152_result.txt', 'a') as file:
        if data_type == 'string':
            logging.info(f'Передача параметру {parameter} текста длиной 150 символов')
            req = t.set_param("{}_{}_{}".format(testing_devcls, testing_slot, parameter), "".rjust(150, '2'))["RACK"][
                "PARAM"]["@Msg"]
            if req == '':
                file.write(
                    f'Вместо ошибки [ {error_06} ] {parameter} принял строку длиной 150 символов\n')
            elif req == error_06:
                logging.info(f'Проверка передачи параметру {parameter} текста длиной 150 символов прошла успешно')
            else:
                file.write(f'При передаче параметру {parameter} строку длиной 150 символов вместо ошибки [ {error_06} ] пришел ответ [ {req} ]\n')
        elif data_type == 'uint' or 'float':
            logging.info(f'Передача параметру {parameter} текста длиной 5 символов')
            req = t.set_param("{}_{}_{}".format(testing_devcls, testing_slot, parameter), "".rjust(5, 'a'))["RACK"][
                "PARAM"]["@Msg"]
            if req == '':
                file.write(
                    f'Вместо ошибки [ {error_42} ] {parameter} принял текстовые данные\n')
            elif req == error_42:
                logging.info(f'Проверка передачи параметру {parameter} текста длиной 5 символов прошла успешно')
            else:
                file.write(f'При передаче параметру {parameter} текстовых данных вместо ошибки [ {error_42} ] пришел ответ [ {req} ]\n')


def hardware_threshold(parameter, HMax, HMin):
    """
    Метод проверяет параметры с аппаратными границами.
    Проверяется соблюдение границ.
    """
    req_hmax = t.set_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter), str(float(HMax) + 1))["RACK"]["PARAM"]["@Msg"]
    req_hmin = t.set_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter), str(float(HMin) - 1))["RACK"]["PARAM"]["@Msg"]
    with open('TEST-T152_result.txt', 'a') as file:
        logging.info(f'Передача параметру {parameter} значение {str(float(HMax) + 1)} выше аппаратной границы')
        if req_hmax == '':
            file.write(f'Вместо ошибки [ {error_04} ] {parameter} принял значение {str(float(HMax) + 1)}\n')
        elif req_hmax == error_04:
            logging.info(f'Проверка передачи параметру {parameter} значения {str(float(HMax) + 1)} выше аппаратной границы прошла успешно')
        else:
            file.write(f'При передаче значения {str(float(HMax) + 1)} {parameter} вместо ошибки [ {error_04} ] пришел ответ [ {req_hmax} ]\n')
        logging.info(
            f'Передача параметру {parameter} значение {str(float(HMin) - 1)} ниже аппаратной границы')
        if req_hmin == '':
            file.write(f'Вместо ошибки [ {error_04} ] {parameter} принял значение {str(float(HMin) - 1)}\n')
        elif req_hmax == error_04:
            logging.info(f'Проверка передачи параметру {parameter} значения {str(float(HMin) - 1)} ниже аппаратной границы прошла успешно')
        else:
            file.write(f'При передаче значения {str(float(HMin) - 1)} {parameter} вместо ошибки [ {error_04} ] пришел ответ [ {req_hmax} ]\n')


def error_threshold(result, parameter, stage):
    """
    Метод, который проверяет параметры на монотонность с границами
    типа warning и critical
    """
    with open('TEST-T152_result.txt', 'a') as file:
        logging.info(f'Параметр {parameter} проверяется на монотонность при {stage}')
        if result == '':
            file.write(f'При проверке {parameter} на монотонность при выставлении {stage} вместо ошибки [ {error_07} ] была нарушена монотонность порогов\n')
        elif result == error_07:
            logging.info(f'Проверка параметра {parameter} на монотонность при {stage} прошла успешно')
        else:
            file.write(f'При проверке {parameter} на монотонность при выставлении {stage} вместо ошибки [ {error_07} ] пришел ответ [ {result} ]\n')


def critical_and_waring_threshold(parameter, CMax, CMin, WMax, WMin):
    """
    Метод проверяет параметры с аппаратными границами.
    Проверяется соблюдение границ.
    """
    thr_pair_value: list[tuple] = [
        ("WMax", float(CMax) + 1), ("WMax", float(CMin) - 1), ("WMax", float(WMin) - 1),
        ("WMin", float(CMax) + 1), ("WMin", float(WMax) + 1), ("WMin", float(CMin) - 1),
        ("CMax", float(WMax) - 1), ("CMax", float(CMin) - 1), ("CMax", float(WMin) - 1),
        ("CMin", float(WMax) + 1), ("CMin", float(CMax) + 1), ("CMin", float(WMin) + 1)
    ]

    req_var: list = [
        'req_WMax_more_CMax', 'req_WMax_less_CMin', 'req_WMax_less_WMin', 'req_WMin_more_CMax',
        'req_WMin_more_WMax', 'req_WMin_less_CMin', 'req_CMax_less_CMax', 'req_CMax_less_CMin',
        'req_CMax_less_WMin', 'req_CMin_more_WMax', 'req_CMin_more_CMax', 'req_CMin_more_WMin'
    ]

    stages: list = [
        'WMax > CMax', 'WMax < CMin', 'WMax < WMin',
        'WMin > CMax', 'WMin > WMax', 'WMin < CMin',
        'CMax < WMax', 'CMax < CMin', 'CMax < WMin',
        'CMin > WMax', 'CMin > CMax', 'CMin > WMin'
    ]

    result_thr_param: dict = dict(zip(req_var, map(lambda item:
                                                   t.set_param(f'{testing_devcls}_{testing_slot}_{parameter + item[0]}',
                                                               str(float('{:.1f}'.format(float(item[1])))))["RACK"][
                                                       "PARAM"]["@Msg"], thr_pair_value)))

    for k, stage in zip(result_thr_param, stages):
        error_threshold(result_thr_param[k], parameter, stage)


def uint_and_float(parameter, data_type):
    """
    Метод, который проверяет параметры с типами данных float и uint.
    Определяет какие границы есть у параметра и исходя из этого
    запускает соответсвующую проверку/
    проверяет на невозможность отправки параметрам текстовые данные
    """
    if "WMax" or "WMin" in parameter:
        parameter = re.sub(r'WM\w\w', '', parameter)
    HMin = t.get_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter + "HMin"))
    HMax = t.get_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter + "HMax"))
    CMin = t.get_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter + "CMin"))
    CMax = t.get_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter + "CMax"))
    WMin = t.get_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter + "WMin"))
    WMax = t.get_param('{}_{}_{}'.format(testing_devcls, testing_slot, parameter + "WMax"))
    if HMin == 'Ключ не найден' and \
            HMax == 'Ключ не найден' and \
            CMin == 'Ключ не найден' and \
            CMax == 'Ключ не найден' and \
            WMin == 'Ключ не найден' and \
            WMax == 'Ключ не найден':
        # числовые параметры без границ, проверить на недопустимость передачи текста
        send_str(parameter, data_type)
    elif HMin != 'Ключ не найден' and HMax != 'Ключ не найден':
        # числовые параметры с аппаратными границами, проверить на соблюдение границ и недопустимость передачи текста
        send_str(parameter, data_type)
        hardware_threshold(parameter, HMax, HMin)
    elif CMin != 'Ключ не найден' and \
            CMax != 'Ключ не найден' and \
            WMin != 'Ключ не найден' and \
            WMax != 'Ключ не найден' and \
            parameter not in deal_params:
        # числовые параметры с границами critical & warning, монотонность границ и недопустимость передачи текста
        send_str(parameter + "WMax", data_type)
        send_str(parameter + "WMin", data_type)
        send_str(parameter + "CMax", data_type)
        send_str(parameter + "CMin", data_type)
        critical_and_waring_threshold(parameter, CMax, CMin, WMax, WMin)
        deal_params.append(parameter)


def something(all_name):
    """
    Метод, который поочередно подбирает для каждого параметра функции для его проверки.
    """
    workbook = openpyxl.load_workbook('profile_onega_vga.xlsx')
    sheet = workbook['parameters']
    for parameter in all_name:
        for row in range(3, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == parameter:
                pass
                data_type = sheet.cell(row=row, column=10).value
                if '.ASM.InPwRef.Set' in parameter:
                    only_numbers(parameter)
                elif '.ASM.InPwRef.AT.Period.Set' in parameter:
                    only_numbers(parameter)
                elif data_type == 'enum':
                    logging.info(f'Параметр {parameter} имеет тип данных выбор - пропуск.')
                elif data_type == 'string':
                    send_str(parameter, data_type)
                elif data_type == 'uint' or 'float':
                    uint_and_float(parameter, data_type)
                break


if __name__ == '__main__':
    IP = input('Введите IP шасси: ')  # 192.168.31.196
    testing_devcls = input('Введите класс тестируемого устройства: ')  # evstc-4.1.0
    testing_slot = input('Введите слот, в котором находится тестируемое устройство: ')  # 6
    t = SessionAPI('s_ap', IP)
    sleep(5)
    name_device = t.get_param(pId.format(testing_devcls, testing_slot))
    logging.info(f'Проверяется усилитель {name_device}')
    logging.info('Сброс параметров в дефолт')
    t.set_param(ConfigReset.format(testing_devcls, testing_slot), '1')
    sleep(30)
    adm_state_state = t.get_param(AdmStateEA.format(testing_devcls, testing_slot, '1'))
    if adm_state_state == '0':
        t.set_param(AdmStateEA.format(testing_devcls, testing_slot, '1'), '1')
        sleep(5)
    elif adm_state_state == '1':
        pass
    else:
        t.set_param(AdmStateEA.format(testing_devcls, testing_slot, '1'), '1')
        sleep(5)
        t.set_param(AdmStateEA.format(testing_devcls, testing_slot, '1'), '0')
        sleep(5)
    something(t.get_param_backup_data(testing_slot, testing_devcls).keys())
    t.logout()
